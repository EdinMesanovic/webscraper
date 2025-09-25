from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from collections import defaultdict
from openpyxl import Workbook
import time
import datetime
import subprocess
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

# --- Konstante ---
# --- Funkcije ---

def init_driver():
    driver = webdriver.Chrome()
    driver.maximize_window()
    return driver

def apply_filters(driver, from_date, to_date):
    WebDriverWait(driver, 10).until(EC.url_contains("/admin/dashboard"))
    driver.get("https://korpa.ba/admin/restaurant_report")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "s_from")))

    start_date_input = driver.find_element(By.ID, "s_from")
    end_date_input = driver.find_element(By.ID, "s_to")

    start_date_input.clear()
    start_date_input.send_keys(from_date)

    end_date_input.clear()
    end_date_input.send_keys(to_date)

    driver.find_element(By.XPATH, '//button[@type="submit"]').click()
    time.sleep(5)

def scrape_orders(driver):
    prva_smjena_orders = []
    druga_smjena_orders = []
    ukupno_prva_smjena = defaultdict(lambda: {"Kolicina": 0, "Cijena": 0.0})
    ukupno_druga_smjena = defaultdict(lambda: {"Kolicina": 0, "Cijena": 0.0})
    ukupno_stavke = defaultdict(lambda: {"Kolicina": 0, "Cijena": 0.0})

    order_links = driver.find_elements(By.CSS_SELECTOR, "div.links a")
    order_urls = []
    for link in order_links:
        url = link.get_attribute("href")
        if url and url.strip() != "" and not url.endswith("/admin/view_order/"):
            order_urls.append(url)

    print(f"Pronađeno {len(order_urls)} validnih narudžbi.")

    for i, order_url in enumerate(order_urls):
        try:
            driver.get(order_url)

            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "my-card")))

            try:
                id_elem = driver.find_element(By.XPATH, '//h3[contains(@class, "content-header-title")]')
                narudzba_id = id_elem.text.strip().split(":")[1].strip()
            except:
                narudzba_id = "NepoznatID"

            try:
                breadcrumb_items = driver.find_elements(By.XPATH, '//ol[contains(@class, "breadcrumb")]/li')
                vrijeme_obj = None
                sat = 0
                for item in breadcrumb_items:
                    text = item.text.strip()
                    if "." in text and ":" in text:
                        vrijeme_obj = datetime.datetime.strptime(text, "%d.%m.%Y %H:%M:%S")
                        sat = vrijeme_obj.hour
                        break
            except:
                sat = 0

            items_container = driver.find_element(By.CLASS_NAME, "my-card")
            items_rows = items_container.find_elements(By.CLASS_NAME, "row")

            stavke = []

            i_row = 0
            while i_row < len(items_rows):
                try:
                    item = items_rows[i_row]
                    item_name_elem = item.find_elements(By.TAG_NAME, "dt")
                    item_price_elem = item.find_elements(By.TAG_NAME, "dd")

                    if not item_name_elem or not item_price_elem:
                        i_row += 1
                        continue

                    main_item_name = item_name_elem[0].text.strip()

                    try:
                        quantity_badge = item_name_elem[0].find_element(By.CLASS_NAME, "badge-success")
                        quantity_text = quantity_badge.text.strip().replace('x', '').strip()
                        quantity = quantity_text
                        main_item_name = main_item_name.replace(quantity_badge.text, '').strip()
                    except:
                        quantity = "1"

                    main_item_price = item_price_elem[0].text.strip().replace('KM', '').strip()

                    dodatak = ""
                    if len(item_name_elem) > 1:
                        dodatak_text = item_name_elem[1].text.strip()
                        if "Triple" in dodatak_text or "Double" in dodatak_text:
                            dodatak = dodatak_text.split("(")[0].strip()

                    final_name = f"{dodatak} {main_item_name}" if dodatak else main_item_name

                    stavke.append({
                        "Naziv": final_name,
                        "Kolicina": quantity,
                        "Cijena": main_item_price
                    })

                    ukupno_stavke[final_name]["Kolicina"] += int(quantity)
                    ukupno_stavke[final_name]["Cijena"] += float(main_item_price) * int(quantity)

                    if sat < 16:
                        ukupno_prva_smjena[final_name]["Kolicina"] += int(quantity)
                        ukupno_prva_smjena[final_name]["Cijena"] += float(main_item_price) * int(quantity)
                    else:
                        ukupno_druga_smjena[final_name]["Kolicina"] += int(quantity)
                        ukupno_druga_smjena[final_name]["Cijena"] += float(main_item_price) * int(quantity)

                except Exception as e:
                    print(f"Greška prilikom čitanja stavke: {e}")

                i_row += 1

            order_data = {
                "id": narudzba_id,
                "stavke": stavke
            }

            if sat < 16:
                prva_smjena_orders.append(order_data)
            else:
                druga_smjena_orders.append(order_data)

            print(f"Uspješno obrađena narudžba {i + 1}/{len(order_urls)}.")

        except Exception as e:
            print(f"Greška kod narudžbe {i + 1}: {e}")
            continue

    return prva_smjena_orders, druga_smjena_orders, ukupno_prva_smjena, ukupno_druga_smjena, ukupno_stavke

def save_to_excel(prva_smjena_orders, druga_smjena_orders, ukupno_prva_smjena, ukupno_druga_smjena, ukupno_stavke):
    wb = Workbook()
    ws = wb.active
    ws.title = "Izvještaj"

    ws.append(["Naziv", "Kolicina", "Cijena"])

    ws.append(["--- PRVA SMJENA ---", "", ""])
    for order in prva_smjena_orders:
        ws.append([f"ID:{order['id']}", "", ""])
        for item in order["stavke"]:
            ws.append([item["Naziv"], item["Kolicina"], item["Cijena"]])
        ws.append([])

    ws.append(["--- UKUPNO PRVA SMJENA ---", "", ""])
    for naziv, podaci in ukupno_prva_smjena.items():
        ws.append([naziv, podaci["Kolicina"], round(podaci["Cijena"], 2)])

    ws.append(["--- DRUGA SMJENA ---", "", ""])
    for order in druga_smjena_orders:
        ws.append([f"ID:{order['id']}", "", ""])
        for item in order["stavke"]:
            ws.append([item["Naziv"], item["Kolicina"], item["Cijena"]])
        ws.append([])

    ws.append(["--- UKUPNO DRUGA SMJENA ---", "", ""])
    for naziv, podaci in ukupno_druga_smjena.items():
        ws.append([naziv, podaci["Kolicina"], round(podaci["Cijena"], 2)])

    ws.append(["--- UKUPNO PRODAJA ---", "", ""])
    for naziv, podaci in ukupno_stavke.items():
        ws.append([naziv, podaci["Kolicina"], round(podaci["Cijena"], 2)])

    filename = f"narudzbe_{datetime.date.today()}.xlsx"

    for col_index, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(col_index)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(filename)

    print(f"\nExcel izvještaj spremljen: {filename} ✅")
    subprocess.Popen(["open", filename])

def main():
    load_dotenv()
    use_env = input("Želite li koristiti podatke iz .env fajla? (y/n): ").strip().lower()
    if use_env == "y":
        email = os.getenv("EMAIL", "")
        password = os.getenv("PASSWORD", "")
        if not email or not password:
            print("Nisu pronađeni EMAIL i PASSWORD u .env fajlu. Unesite ručno.")
            email = input("Unesi email: ").strip()
            password = input("Unesi šifru: ").strip()
    else:
        email = input("Unesi email: ").strip()
        password = input("Unesi šifru: ").strip()
    driver = init_driver()
    try:
        login_with_credentials(driver, email, password)

        use_test_date = input("Da li želiš koristiti testni datum? (y/n): ").strip().lower()
        if use_test_date == "y":
            from_date = "29/04/2025"
            to_date = "29/04/2025"
            print(f"Koristi se testni datum: {from_date} do {to_date}")
        else:
            from_date = input("Unesi OD datum (u formatu DD/MM/YYYY) ili Enter za danas: ").strip() or datetime.date.today().strftime("%m/%d/%Y")
            to_date = input("Unesi DO datum (u formatu DD/MM/YYYY) ili Enter za danas: ").strip() or datetime.date.today().strftime("%m/%d/%Y")

        apply_filters(driver, from_date, to_date)
        prva_smjena_orders, druga_smjena_orders, ukupno_prva_smjena, ukupno_druga_smjena, ukupno_stavke = scrape_orders(driver)
        save_to_excel(prva_smjena_orders, druga_smjena_orders, ukupno_prva_smjena, ukupno_druga_smjena, ukupno_stavke)
    finally:
        driver.quit()

def login_with_credentials(driver, email, password):
    driver.get("https://korpa.ba/admin")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "email")))
    email_input = driver.find_element(By.NAME, "email")
    password_input = driver.find_element(By.NAME, "password")
    email_input.send_keys(email)
    password_input.send_keys(password)
    driver.find_element(By.XPATH, '//button[@type="submit"]').click()

# --- Pokreni program ---
if __name__ == "__main__":
    main()
